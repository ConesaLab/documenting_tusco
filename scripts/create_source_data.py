#!/usr/bin/env python3
"""Create Source_Data.xlsx for Nature Communications submission.

One sheet per figure panel with aggregated/plotted values from existing TSVs.
"""

import csv
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

LENGTH_BINS = [
    ("<600 bp", 0, 600),
    ("600 bp\u20131 kb", 600, 1000),
    ("1\u20132 kb", 1000, 2000),
    ("2\u20134 kb", 2000, 4000),
    (">4 kb", 4000, float("inf")),
]


def read_tsv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def clean_val(v):
    """Convert 'NA' strings to None (empty cell), parse numbers."""
    if v is None or v == "" or v == "NA":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return v


def add_sheet(wb, sheet_name, headers, rows, description=""):
    ws = wb.create_sheet(title=sheet_name)
    start_row = 1
    if description:
        ws.cell(row=1, column=1, value=description).font = Font(italic=True)
        start_row = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h).font = Font(bold=True)
    for r, row in enumerate(rows, start_row + 1):
        for c, val in enumerate(row, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=5)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def fig2ab(wb):
    """Fig 2a-b: Tissue distribution of TUSCO genes."""
    rows = []
    for species, fname in [
        ("Human", "data/processed/tusco/hsa/tusco_human_tissue_statistics.tsv"),
        ("Mouse", "data/processed/tusco/mmu/tusco_mouse_tissue_statistics.tsv"),
    ]:
        data = read_tsv(os.path.join(REPO, fname))
        for row in data:
            rows.append([
                species,
                row["Anatomical_Entity_Name"],
                int(row["Number_of_TUSCO_Genes"]),
            ])
    add_sheet(wb, "Fig 2a-b", ["Species", "Tissue", "Number of TUSCO genes"], rows,
              "Tissue distribution of genes in the TUSCO gene set (human and mouse).")


def fig2c(wb):
    """Fig 2c: Transcript length distributions binned."""
    data = read_tsv(os.path.join(REPO, "figs/figure-02/tables/fig-2c.tsv"))
    counts = defaultdict(lambda: defaultdict(int))
    for row in data:
        if not row.get("transcript_length"):
            continue
        length = float(row["transcript_length"])
        dataset = row.get("dataset", "")
        for label, lo, hi in LENGTH_BINS:
            if lo <= length < hi:
                counts[dataset][label] += 1
                break
    bin_labels = [b[0] for b in LENGTH_BINS]
    rows = []
    for dataset in sorted(counts.keys()):
        for bl in bin_labels:
            rows.append([dataset, bl, counts[dataset].get(bl, 0)])
    add_sheet(wb, "Fig 2c", ["Dataset", "Length bin", "Count"], rows,
              "Transcript length distributions by dataset (binned counts).")


def fig2d(wb):
    """Fig 2d: Cross-tissue median expression."""
    data = read_tsv(os.path.join(REPO, "figs/figure-02/tables/fig-2d.tsv"))
    headers = ["Species", "Group", "Median value", "Log10 median value"]
    rows = []
    for row in data:
        rows.append([
            clean_val(row.get("Species")),
            clean_val(row.get("Group")),
            clean_val(row.get("median_value")),
            clean_val(row.get("log_median_value")),
        ])
    add_sheet(wb, "Fig 2d", headers, rows,
              "Cross-tissue median expression values (log10 TPM).")


def fig3a(wb):
    """Fig 3a: TUSCO vs SIRV comparison — dumbbell + bar data."""
    # Dumbbell plot data (radar_metrics)
    dumbbell_rows = []
    for species, fname in [("Human", "figure3a-human.tsv"), ("Mouse", "figure3a-mouse.tsv")]:
        for row in read_tsv(os.path.join(REPO, "figs/figure-03/tables", fname)):
            if row.get("record_type") == "radar_metrics":
                dumbbell_rows.append([
                    species,
                    clean_val(row.get("pipeline")),
                    clean_val(row.get("metric")),
                    clean_val(row.get("sirv")),
                    clean_val(row.get("tusco")),
                ])
    add_sheet(wb, "Fig 3a dumbbell",
              ["Species", "Pipeline", "Metric", "SIRV value", "TUSCO value"],
              dumbbell_rows, "Dumbbell plot: TUSCO vs SIRV metrics per pipeline.")

    # Bar chart data (bar_distribution)
    bar_rows = []
    for species, fname in [("Human", "figure3a-human.tsv"), ("Mouse", "figure3a-mouse.tsv")]:
        for row in read_tsv(os.path.join(REPO, "figs/figure-03/tables", fname)):
            if row.get("record_type") == "bar_distribution":
                bar_rows.append([
                    species,
                    clean_val(row.get("big_category")),
                    clean_val(row.get("final_label")),
                    clean_val(row.get("count")),
                    clean_val(row.get("percentage")),
                    clean_val(row.get("Type")),
                    clean_val(row.get("pipeline")),
                ])
    add_sheet(wb, "Fig 3a bars",
              ["Species", "Category", "Label", "Count", "Percentage", "Type", "Pipeline"],
              bar_rows, "Bar chart: TP/PTP/FP/FN distribution per pipeline.")


def fig3b(wb):
    """Fig 3b: TP/PTP/FP/FN proportions with stats."""
    rows_all = []
    for species, fname in [("Human", "figure3b-human.tsv"), ("Mouse", "figure3b-mouse.tsv")]:
        for row in read_tsv(os.path.join(REPO, "figs/figure-03/tables", fname)):
            rows_all.append([
                species,
                clean_val(row.get("big_category")),
                clean_val(row.get("count")),
                clean_val(row.get("percentage")),
                clean_val(row.get("Type")),
                clean_val(row.get("pipeline")),
                clean_val(row.get("n")),
                clean_val(row.get("record_type")),
                clean_val(row.get("mean_perc")),
                clean_val(row.get("sd_perc")),
                clean_val(row.get("test_method")),
                clean_val(row.get("p_value")),
            ])
    add_sheet(wb, "Fig 3b",
              ["Species", "Category", "Count", "Percentage", "Type", "Pipeline",
               "n", "Record type", "Mean %", "SD %", "Test method", "p-value"],
              rows_all, "TP/PTP/FP/FN proportions with statistical tests.")


def fig3c(wb):
    """Fig 3c: RIN correlation data."""
    data = read_tsv(os.path.join(REPO, "figs/figure-03/tables/figure3c.tsv"))
    rows = []
    for row in data:
        rows.append([
            clean_val(row.get("RIN")),
            clean_val(row.get("TP")),
            clean_val(row.get("PTP")),
            clean_val(row.get("FP")),
            clean_val(row.get("FN")),
            clean_val(row.get("Group")),
            clean_val(row.get("Dataset")),
            clean_val(row.get("TP_TPPTP")),
        ])
    add_sheet(wb, "Fig 3c",
              ["RIN", "TP", "PTP", "FP", "FN", "Sample", "Dataset", "TP/(TP+PTP)"],
              rows, "RNA degradation: RIN vs fraction of fully reconstructed transcripts.")


def fig3d(wb):
    """Fig 3d: FN genes across datasets."""
    rows = []
    for species, fname in [("Human", "figure3d-human.tsv"), ("Mouse", "figure3d-mouse.tsv")]:
        for row in read_tsv(os.path.join(REPO, "figs/figure-03/tables", fname)):
            rows.append([
                species,
                clean_val(row.get("num_samples_shared")),
                clean_val(row.get("num_genes")),
            ])
    add_sheet(wb, "Fig 3d",
              ["Species", "Number of datasets shared", "Number of genes"],
              rows, "FN genes in the TUSCO gene set by number of datasets where detected.")


def fig4b(wb):
    """Fig 4b: TUSCO Novel benchmark."""
    data = read_tsv(os.path.join(REPO, "figs/figure-04/tables/fig-4b.tsv"))
    rows = []
    for row in data:
        rows.append([
            clean_val(row.get("pipeline_label")),
            clean_val(row.get("sample")),
            clean_val(row.get("species")),
            clean_val(row.get("eval_type")),
            clean_val(row.get("Sn")),
            clean_val(row.get("nrPre")),
            clean_val(row.get("1_red")),
            clean_val(row.get("1_FDR")),
            clean_val(row.get("PDR")),
            clean_val(row.get("rPre")),
        ])
    add_sheet(wb, "Fig 4b",
              ["Tool", "Sample", "Species", "Evaluation type",
               "Sensitivity", "nrPrecision", "1/Redundancy", "1-FDR", "PDR", "rPrecision"],
              rows, "TUSCO Novel benchmark: native vs novel reference performance.")


def fig5bc(wb):
    """Fig 5b-c: Replication gains and universal vs tissue-specific."""
    data = read_tsv(os.path.join(REPO, "figs/figure-05/tables/fig-5b-5c.tsv"))

    # Fig 5b: individual points and bar summaries
    rows_5b = []
    for row in data:
        pid = row.get("panel_id", "")
        if pid in ("5b_points", "5b_bars"):
            rows_5b.append([
                clean_val(row.get("Tissue")),
                clean_val(row.get("Samples")),
                clean_val(row.get("Combo")),
                clean_val(row.get("Metric")),
                clean_val(row.get("Value")),
                clean_val(row.get("N")),
                clean_val(row.get("Mean")),
                clean_val(row.get("SD")),
                clean_val(row.get("SE")),
                clean_val(row.get("CI_lower")),
                clean_val(row.get("CI_upper")),
            ])
    add_sheet(wb, "Fig 5b",
              ["Tissue", "Replicates", "Combination", "Metric", "Value",
               "N", "Mean", "SD", "SE", "CI lower", "CI upper"],
              rows_5b, "Replication gains: performance vs replicate number.")

    # Fig 5c: universal vs tissue-specific comparison
    rows_5c = []
    for row in data:
        pid = row.get("panel_id", "")
        if pid in ("5c_points", "5c_summary"):
            rows_5c.append([
                clean_val(row.get("Tissue")),
                clean_val(row.get("Type")),
                clean_val(row.get("Sample")),
                clean_val(row.get("Sn")),
                clean_val(row.get("nrPre")),
                clean_val(row.get("1/red")),
                clean_val(row.get("1-FDR")),
                clean_val(row.get("PDR")),
                clean_val(row.get("rPre")),
            ])
    add_sheet(wb, "Fig 5c",
              ["Tissue", "Gene set type", "Sample", "Sensitivity", "nrPrecision",
               "1/Redundancy", "1-FDR", "PDR", "rPrecision"],
              rows_5c, "Universal vs tissue-specific TUSCO gene sets.")


def main():
    wb = Workbook()
    wb.remove(wb.active)

    fig2ab(wb)
    fig2c(wb)
    fig2d(wb)
    fig3a(wb)  # creates two sheets: "Fig 3a dumbbell" and "Fig 3a bars"
    fig3b(wb)
    fig3c(wb)
    fig3d(wb)
    fig4b(wb)
    fig5bc(wb)  # creates two sheets: "Fig 5b" and "Fig 5c"

    out_path = os.path.join(REPO, "submission/01_for_upload/Source_Data.xlsx")
    wb.save(out_path)
    print(f"Created {out_path}")
    for ws in wb.worksheets:
        print(f"  Sheet '{ws.title}': {ws.max_row} total rows")


if __name__ == "__main__":
    main()
